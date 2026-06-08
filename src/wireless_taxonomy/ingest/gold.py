from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from wireless_taxonomy.textnorm import normalize_doi, normalize_title


@dataclass(frozen=True)
class GoldRecord:
    title: str
    venue: str
    year: int
    doi: str | None
    normalized_title: str
    normalized_doi: str
    raw: dict[str, Any] = field(default_factory=dict)


_TITLE_KEYS = ("paper title", "title", "name", "paper")
_DOI_KEYS = ("doi", "doi version of key", "doi url")
_VENUE_KEYS = ("conference", "venue", "conf")
_YEAR_KEYS = ("year", "yr")
_WIRELESS_KEYS = ("wireless", "is_wireless", "wireless?", "wireless candidate")
_TRUE_VALUES = {"1", "yes", "y", "true", "t", "wireless", "x"}


class GoldSheetReader:
    """Reads a manually curated gold sheet (csv/xlsx) into GoldRecords.

    CSV is parsed with the stdlib; xlsx is parsed lazily via ``openpyxl`` (an
    optional dependency). Column names are matched case-insensitively against
    common variants so the user's own sheet layout need not be reshaped first.
    """

    def __init__(
        self,
        path: str,
        default_venue: str | None = None,
        default_year: int | None = None,
        wireless_only: bool = False,
    ) -> None:
        self.path = Path(path)
        self.default_venue = default_venue
        self.default_year = default_year
        self.wireless_only = wireless_only

    def read(self) -> list[GoldRecord]:
        columns, raw_rows = self._load_rows()
        columns = [str(c).strip() for c in columns]
        lower_map = {c.lower(): c for c in columns}

        title_col = _find(lower_map, _TITLE_KEYS)
        if title_col is None:
            raise ValueError(
                f"Could not find a title column in {self.path.name}. "
                f"Columns seen: {columns}"
            )
        doi_col = _find(lower_map, _DOI_KEYS)
        venue_col = _find(lower_map, _VENUE_KEYS)
        year_col = _find(lower_map, _YEAR_KEYS)
        wireless_col = _find(lower_map, _WIRELESS_KEYS)

        records: list[GoldRecord] = []
        for raw_row in raw_rows:
            title = _clean(raw_row.get(title_col))
            if not title:
                continue
            if self.wireless_only and wireless_col is not None:
                if _clean(raw_row.get(wireless_col)).lower() not in _TRUE_VALUES:
                    continue
            venue = _clean(raw_row.get(venue_col)) if venue_col else ""
            venue = venue or (self.default_venue or "")
            year = _coerce_year(raw_row.get(year_col) if year_col else None, self.default_year)
            if not venue or year is None:
                raise ValueError(
                    f"Row '{title[:60]}' is missing venue/year and no defaults were provided. "
                    "Pass --venue/--year or include conference/year columns."
                )
            doi = _clean(raw_row.get(doi_col)) if doi_col else ""
            records.append(
                GoldRecord(
                    title=title,
                    venue=venue,
                    year=year,
                    doi=doi or None,
                    normalized_title=normalize_title(title),
                    normalized_doi=normalize_doi(doi),
                    raw={str(k): _jsonable(v) for k, v in raw_row.items()},
                )
            )
        return records

    def _load_rows(self) -> tuple[list[str], list[dict[str, Any]]]:
        if self.path.suffix.lower() in {".xlsx", ".xls"}:
            return self._load_xlsx()
        return self._load_csv()

    def _load_csv(self) -> tuple[list[str], list[dict[str, Any]]]:
        import csv as _csv

        with self.path.open(newline="", encoding="utf-8-sig") as fh:
            reader = _csv.DictReader(fh)
            columns = [str(c).strip() for c in (reader.fieldnames or [])]
            rows: list[dict[str, Any]] = []
            for row in reader:
                rows.append({str(k).strip(): v for k, v in row.items() if k is not None})
        return columns, rows

    def _load_xlsx(self) -> tuple[list[str], list[dict[str, Any]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - optional dependency
            raise RuntimeError(
                "Reading .xlsx gold sheets requires openpyxl. Install it with "
                "`pip install openpyxl`, or export the sheet to CSV."
            ) from exc

        workbook = load_workbook(self.path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return [], []
        columns = [str(c).strip() if c is not None else "" for c in header]
        rows: list[dict[str, Any]] = []
        for values in rows_iter:
            rows.append({columns[i]: values[i] for i in range(min(len(columns), len(values)))})
        workbook.close()
        return columns, rows


def distinct_venue_years(
    paths: list[str],
    default_venue: str | None = None,
    default_year: int | None = None,
) -> list[tuple[str, int]]:
    """Return the sorted, de-duplicated (venue, year) pairs across gold sheet(s).

    Lets the eval harness drive itself off whatever conferences a dropped-in
    sheet actually contains, instead of a hardcoded venue list.
    """
    seen: dict[tuple[str, int], None] = {}
    for path in paths:
        for record in GoldSheetReader(path, default_venue, default_year).read():
            seen.setdefault((record.venue, record.year), None)
    return sorted(seen, key=lambda vy: (vy[0].lower(), vy[1]))


def _find(lower_map: dict[str, str], keys: tuple[str, ...]) -> str | None:
    for key in keys:
        if key in lower_map:
            return lower_map[key]
    return None


def _clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    return text


def _coerce_year(value: Any, default: int | None) -> int | None:
    text = _clean(value)
    if text:
        try:
            return int(float(text))
        except (TypeError, ValueError):
            return default
    return default


def _jsonable(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)
